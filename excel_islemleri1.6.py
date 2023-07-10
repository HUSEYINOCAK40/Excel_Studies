import tkinter as tk
from tkinter import filedialog
import os
import xlrd
import openpyxl
import xlwt

def select_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def convert_to_xls(file_path):
    workbook = openpyxl.load_workbook(file_path)
    converted_file_path = os.path.splitext(file_path)[0] + ".xls"
    converted_workbook = xlwt.Workbook()

    for sheet_name in workbook.sheetnames:
        sheet_name = sheet_name[:31]
        converted_sheet = converted_workbook.add_sheet(sheet_name)
        worksheet = workbook[sheet_name]
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                converted_sheet.write(row - 1, col - 1, cell_value)

    converted_workbook.save(converted_file_path)
    return converted_file_path

def merge_sheets_into_one():
    # Kullanıcının seçtiği klasör yolunu alır
    folder_path = folder_entry.get()

    # Çıktı dosya adını oluşturur
    output_file = os.path.join(folder_path, "işlenmiş_tek_sheet.xls")

    # Dosya adında numaralandırma için kullanılan sayaç
    counter = 1
    new_output_file = output_file

    # Dosya adı mevcut olduğu sürece numaralandırma yapar
    while os.path.exists(new_output_file):
        filename, extension = os.path.splitext(output_file)
        new_output_file = f"{filename}({counter}){extension}"
        counter += 1

    # Klasördeki tüm dosyaları alır
    files = [file for file in os.listdir(folder_path) if file.endswith('.xls') or file.endswith('.xlsx')]

    # Çıktı Workbook'ü oluşturur
    output_workbook = xlwt.Workbook()
    output_sheet = output_workbook.add_sheet("Sheet1")

    # Satır indeksini başlatır
    row_index = 0

    # Dönüştürülen XLS dosyalarını tutmak için liste oluşturur
    converted_files = []

    # Dosya listesini döngüye alır
    for file in files:
        file_path = os.path.join(folder_path, file)

        # Dosyanın uzantısına göre işlem yapar
        if file.endswith('.xls'):
            # XLS dosyasını yükler
            workbook = xlrd.open_workbook(file_path)
        else:
            # XLSX dosyasını XLS formatına dönüştürür
            file_path = convert_to_xls(file_path)
            workbook = xlrd.open_workbook(file_path)

            # Dönüştürülen XLS dosyasını listeye ekler
            converted_files.append(file_path)

        # Dosyanın içindeki her bir sayfayı döngüye alır
        for sheet_name in workbook.sheet_names():
            # Sayfa adını alır
            worksheet = workbook.sheet_by_name(sheet_name)

            # Sayfa içindeki her bir satırı döngüye alır
            for row in range(worksheet.nrows):
                # Satırda boş olmayan bir hücre varsa işlem yapar
                if any(cell_value for cell_value in worksheet.row_values(row)):
                    # Satırdaki her bir hücreyi kopyalar
                    for col, cell_value in enumerate(worksheet.row_values(row)):
                        output_sheet.write(row_index, col, cell_value)

                    # Satır indeksini artırır
                    row_index += 1

    # Çıktı dosyasını kaydeder
    output_workbook.save(new_output_file)

    # Dönüştürülen XLS dosyalarını siler
    for converted_file in converted_files:
        os.remove(converted_file)

    # İşlem tamamlandı bilgisini gösterir
    tk.messagebox.showinfo("Bilgi", f"Excel dosyaları tek bir çalışma sayfasında birleştirildi.\nÇıktı dosyası: {new_output_file}")


def merge_excel_files():
    folder_path = folder_entry.get()
    xls_files = [file for file in os.listdir(folder_path) if file.endswith('.xls')]
    xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
    output_workbook = xlwt.Workbook()
        # Dosya adında numaralandırma için kullanılan sayaç
        
    # XLS dosyalarını birleştir
    for file in xls_files:
        file_path = os.path.join(folder_path, file)
        workbook = xlrd.open_workbook(file_path)
        for sheet_index, sheet_name in enumerate(workbook.sheet_names()):
            unique_sheet_name = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_index + 1}"
            output_sheet = output_workbook.add_sheet(unique_sheet_name)
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row, col)
                    output_sheet.write(row, col, cell_value)
    
    # XLSX dosyalarını XLS formatına dönüştürerek birleştir
    for file in xlsx_files:
        file_path = os.path.join(folder_path, file)
        converted_file_path = convert_to_xls(file_path)
        workbook = xlrd.open_workbook(converted_file_path)
        for sheet_index, sheet_name in enumerate(workbook.sheet_names()):
            unique_sheet_name = f"{os.path.splitext(os.path.basename(converted_file_path))[0]}_{sheet_index + 1}"
            output_sheet = output_workbook.add_sheet(unique_sheet_name)
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row, col)
                    output_sheet.write(row, col, cell_value)
        
        # Dönüştürülen XLS dosyasını sil
        os.remove(converted_file_path)

    # Çıktı dosyasını kaydet
    output_file = os.path.join(folder_path, "işlenmiş_dosya.xls")

    counter = 1
    new_output_file = output_file

    # Dosya adı mevcut olduğu sürece numaralandırma yapar
    while os.path.exists(new_output_file):
        filename, extension = os.path.splitext(output_file)
        new_output_file = f"{filename}({counter}){extension}"
        counter += 1

    output_workbook.save(new_output_file)
    # İşlem tamamlandı bilgisi
    tk.messagebox.showinfo("Bilgi", f"Excel dosyaları tek bir dosyada birleştirildi.\nÇıktı dosyası: {new_output_file}")

def create_unique_folder(target_folder_path):
    # Hedef klasörü oluşturur veya numaralandırır
    folder_number = 1
    while os.path.exists(target_folder_path):
        target_folder_path = f"{target_folder_path}_{folder_number}"
        folder_number += 1

    # Klasörü oluşturur
    os.makedirs(target_folder_path)

    return target_folder_path

def separate_excel_sheets():
    # Seçilen dosya yolunu alır
    folder_path = folder_entry.get()

    # Hedef klasör adını oluşturur
    target_folder = "ayrılmış dosyalar"

    # Hedef klasörün tam yolunu oluşturur
    target_folder_path = os.path.join(folder_path, target_folder)

    # Hedef klasörü oluşturur veya numaralandırır
    target_folder_path = create_unique_folder(target_folder_path)

    # Klasördeki bütün XLS dosyalarını alır
    xls_files = [file for file in os.listdir(folder_path) if file.endswith('.xls')]

    # XLS dosyalarını işle
    for xls_file in xls_files:
        xls_file_path = os.path.join(folder_path, xls_file)
        workbook = xlrd.open_workbook(xls_file_path)
        for sheet_name in workbook.sheet_names():
            # Hedef dosyanın adını ve sheet adını oluşturur
            output_file_name = f"{os.path.splitext(xls_file)[0]}_{sheet_name}.xls"
            output_file_path = os.path.join(target_folder_path, output_file_name)

            # Yeni XLS dosyasını oluşturur
            output_workbook = xlwt.Workbook()
            output_sheet = output_workbook.add_sheet(sheet_name)

            # Verileri kopyalar
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row, col)
                    output_sheet.write(row, col, cell_value)

            # Yeni XLS dosyasını kaydeder
            output_workbook.save(output_file_path)

    # XLSX dosyalarını XLS formatına dönüştürür ve işle
    xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
    converted_files = []
    for xlsx_file in xlsx_files:
        xlsx_file_path = os.path.join(folder_path, xlsx_file)
        converted_file_path = convert_to_xls(xlsx_file_path)
        converted_files.append(converted_file_path)
        workbook = xlrd.open_workbook(converted_file_path)
        for sheet_name in workbook.sheet_names():
            # Hedef dosyanın adını ve sheet adını oluşturur
            output_file_name = f"{os.path.splitext(xlsx_file)[0]}_{sheet_name}.xls"
            output_file_path = os.path.join(target_folder_path, output_file_name)

            # Yeni XLS dosyasını oluşturur
            output_workbook = xlwt.Workbook()
            output_sheet = output_workbook.add_sheet(sheet_name)

            # Verileri kopyalar
            worksheet = workbook.sheet_by_name(sheet_name)
            for row in range(worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row, col)
                    output_sheet.write(row, col, cell_value)

            # Yeni XLS dosyasını kaydeder
            output_workbook.save(output_file_path)

    # Dönüştürülen XLS dosyalarını siler
    for converted_file in converted_files:
        os.remove(converted_file)

    # İşlem tamamlandı bilgisini gösterir
    tk.messagebox.showinfo("Bilgi", "Excel dosyaları çalışma sayfalarına bölünerek 'Ayrılmış Klasör'e kaydedildi.")


window = tk.Tk()
window.title("Excel İşlemleri(XLS ve XLSX Dosyaları)")
window.geometry("400x150")

folder_frame = tk.Frame(window)
folder_frame.pack(pady=20)

folder_label = tk.Label(folder_frame, text="Klasör Seç:")
folder_label.pack(side=tk.LEFT)

folder_entry = tk.Entry(folder_frame, width=30)
folder_entry.pack(side=tk.LEFT)

merge_sheets_button = tk.Button(window, text="Dosyaları Tek Sheette Topla", command=merge_sheets_into_one)
merge_sheets_button.pack()

select_folder_button = tk.Button(folder_frame, text="Klasör Seç", command=select_folder)
select_folder_button.pack(side=tk.LEFT)

merge_button = tk.Button(window, text="Dosyaları Tek Dosyada Ayrı Sheetlerde Topla", command=merge_excel_files)
merge_button.pack()

merge_button = tk.Button(window, text="Dosyaları Sheetlerine Ayır ve Farklı Klasöre Kaydet", command=separate_excel_sheets)
merge_button.pack()

window.mainloop()






